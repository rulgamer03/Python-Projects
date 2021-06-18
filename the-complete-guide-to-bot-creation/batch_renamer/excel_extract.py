import os

from openpyxl import load_workbook, Workbook # load_workbook open already existing books, Workbook create work bools 
from openpyxl.styles import Font # to bolfing text

# mapping from index to letter
letters = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", 
    "J", "K", "L", "M", "N", "O", "P", "Q", "R", 
    "S", "T", "U", "V", "W", "X", "Y", "Z"
]

# needed information about excel workbook
summary_wb_name = "summary.xlsx"
sheet_name = "Source Data"
column = "C"
start_row = 4

# get all the files from the current folder
dir_content = os.listdir(".")
excel_tl_files = [doc for doc in dir_content if doc.endswith("xltx")]
# sorts the names by year
excel_tl_files.sort()
processed = 0

# set up the output workbook
summary_wb = Workbook() # New excel file
summary_ws = summary_wb.active
summary_ws["A1"] = "Customers"
summary_ws["A1"].font = Font(size="16", bold=True)

for index, excel_file in enumerate(excel_tl_files): #enumerate to give a index
    print(f"Processing and extracting data from {excel_file}")
    wb = load_workbook(filename=excel_file)
    sheet = wb[sheet_name]

    # set up the counter and get the current cell
    curr_row = start_row
    #print(curr_row) #4
    cell_num = f"{column}{curr_row}" #column letter and curr row
    #print(cell_num) #C4
    cell = sheet[cell_num] # 
    # print(cell) #<Cell 'Source Data'.C4>


    # empty list of entities for this workbook
    workbook_entities = []

    # get distinct customers from all years
    while cell.value is not None: # cell is empty
        # print(cell.value) Anton bergs ...
        # get the value of the current cell
        entity = cell.value

        # if the entity is not yet part of the array, add it
        if entity not in workbook_entities:
            workbook_entities.append(entity)

        # set the current cell to the next row
        curr_row += 1
        cell_num = f"{column}{curr_row}"
        cell = sheet[cell_num]

    # write to new excel sheet with unique customers and heading of the given year
    column_letter = letters[index] # index by enumerate method
    column_index = f"{column_letter}2" #A2 B2 and ...
    # print(column_index)
    summary_ws.column_dimensions[column_letter].width = 20 #dimention
    # write the name of the file without filetype
    summary_ws[column_index] = os.path.splitext(excel_file)[0] # sales_2011 and ...
    # print(os.path.splitext(excel_file)[0]) #
    summary_ws[column_index].font = Font(bold=True)

    # for every entity, write it to a row below the header
    #print(workbook_entities)
    #['ANTON', 'BERGS', 'BOLID', 'BOTTM', 'ERNSH', 'GODOS', 'HUNGC', 'PICCO', 'RATTC', 'REGGC', 'SAVEA', 'SEVES', 'WHITC', 'ALFKI', 'LINOD', 'QUICK', 'VAFFE', 'BONAP', 'BSBEV', 'FRANS', 'HILAA', 'LAZYK', 'LEHMS', 'MAGAA', 'OTTIK', 'PERIC', 'QUEEN', 'RANCH', 'TRAIH', 'ANATR', 'AROUT', 'CHOPS', 'FAMIA', 'FRANK', 'FURIB', 'GOURL', 'MEREP', 'RICAR', 'RICSU', 'WARTH', 'WOLZA', 'EASTC', 'FOLKO', 'TRADH', 'THEBI', 'BLONP', 'DUMON', 'LAUGB', 'NORTS', 'OLDWO', 'TOMSP', 'VINET', 'CACTU', 'HANAR', 'HUNGO', 'SUPRD', 'TORTU', 'WANDK', 'KOENE', 'MAISD', 'WELLI', 'WILMK', 'LONEP', 'THECR', 'VICTE', 'FRANR', 'LAMAI', 'CONSH', 'GREAL', 'ISLAT', 'MORGK', 'SIMOB', 'BLAUS', 'FOLIG', 'LILAS', 'OCEAN', 'PRINI', 'QUEDE']
    # Enumerate() method adds a counter to an iterable and returns it in a form of enumerate object. 
    for i, entity in enumerate(workbook_entities):
        cell_index = f"{column_letter}{i + 3}"
        # print(cell_index) #A3 A4...  B3 B4... C3 C4...
        summary_ws[cell_index] = entity

    processed += 1

# save the workbook
summary_wb.save(summary_wb_name)
print(f"Processed {processed} of {len(excel_tl_files)} excel files.")
