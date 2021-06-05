import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

def automate_excel(file_name):
    """The file name should have the following structure: sales_month.xlsx"""
    # read excel file
    excel_file = pd.read_excel(file_name)
    # make pivot table
    report_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    # The file has many columns but we’ll only use the Gender, Product line, and Total columns for the report we’re going to create
    # splitting the month and extension from the file name
    month_and_extension = file_name.split('_')[1]
    month_and_extension = month_and_extension + '.xlsx'
    print(month_and_extension) #september.xlsx
    # send the report table to excel file
    report_table.to_excel(f'report_{month_and_extension}', sheet_name='Report', startrow=4)
    # loading workbook and selecting sheet
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report'] # Sheet Report
    # cell references (original spreadsheet)
    min_column = wb.active.min_column
    print(f'min_column {min_column}')
    max_column = wb.active.max_column
    print(f'max_column {wb.active.max_column}')
    min_row = wb.active.min_row
    print(f'min_row {wb.active.min_row}')
    max_row = wb.active.max_row
    print(f'max_row {wb.active.max_row}')
    # adding a chart
    barchart = BarChart()
    data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) #including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) #not including headers
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, "B12") #location chart
    barchart.title = 'Sales by Product line'
    barchart.style = 2 #choose the chart style
    # applying formulas
    # first create alphabet list as references for cells
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] #note: Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    # sum in columns B-G
    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
    # getting month name
    month_name = month_and_extension.split('.')[0]
    # formatting the report
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
    wb.save(f'report_{month_and_extension}')
    return 

automate_excel('sales_september_2021.xlsx')
